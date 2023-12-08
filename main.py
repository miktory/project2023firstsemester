import re
import openpyxl


class Project:
    def __init__(self, title, description, purpose, results, technology, stack):
        self.title = title
        self.description = description
        self.purpose = purpose
        self.results = results
        self.technology = technology
        self.stack = stack
        self.coincidences = 0


class Role:
    def __init__(self, title, keywords):
        self.title = title
        self.keywords = keywords


def find_coincidences(text, words):
    count = 0
    i = -1  # начинаем поиск с начала строки
    for word in words:
        while True:
            # находим следующее вхождение подстроки
            i = text.lower().find(word.lower(), i+1)
            if i == -1:
                break
            count += 1
    return count


def SortByRole(role):
    path = "data.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    projects = []
    for i in range(2,sheet_obj.max_row):
        projects.append(Project(sheet_obj.cell(row = i, column = 1).value, sheet_obj.cell(row = i, column = 2).value, sheet_obj.cell(row = i, column = 3).value, sheet_obj.cell(row = i, column = 4).value,
                     sheet_obj.cell(row = i, column = 5).value, sheet_obj.cell(row = i, column = 6).value))

    for i in range(0, len(projects)):
        info = ' '.join([projects[i].title, projects[i].description, projects[i].purpose, projects[i].results,
                         projects[i].technology, projects[i].stack])
        projects[i].coincidences = find_coincidences(info, role.keywords)

    projects.sort(key=lambda x: x.coincidences)
    projects.reverse()

    for i in range(0, len(projects)):
        print("----", projects[i].title, "----")
        print("# Описание: ", projects[i].description)
        print("# Цель: ", projects[i].purpose)
        print("# Ожидаемый результат: ", projects[i].results)
        print("# Ключевые технологии: ", projects[i].technology)
        print("# Технологический стек: ", projects[i].stack)
        print("# Совпадения: ", projects[i].coincidences)

# Аналитик #
analyst = Role("Аналитик", ["аналитик", "таблиц", "pandas", "пандас", "обучени", "анализ", "sql", "excel", "ml", "статистик", "numpy", "прогноз", "бд", "баз", "данны"])
# Дизайнер #
designer = Role("Дизайнер", ["фронт", "фигм", "figma", "ui", "HTML", "css", "дизайн", "вёрстк", "интерфейс", "illustrator", "photoshop"])
# Бэкенд #
backend = Role("Бэкенд", ["бэк", "код", "API", "разработ", "c#", "c++", "Python", "deploy", "депло", "контейнер", "rest", "парс", "апи", "докер", "docker"])
SortByRole(backend)